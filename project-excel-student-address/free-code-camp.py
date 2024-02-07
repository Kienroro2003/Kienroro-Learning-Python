# import openpyxl
# import xlrd
#
# theFile = openpyxl.load_workbook('/Users/kienroro2003/Downloads/kienroro.xlsx')
# allSheetNames = theFile.sheetnames
#
# print("All sheet names {} " .format(theFile.sheetnames))
#
#
# for x in allSheetNames:
#     print("Current sheet name is {}" .format(x))
#     currentSheet = theFile[x]
#     print(currentSheet['B4'].value)


import openpyxl

url = '/Users/kienroro2003/Downloads/Filemau-C1-HocSinhDiaPhuong-ver2-data (4)KHOI 3 Backup.xlsx'

theFile = openpyxl.load_workbook(url)
allSheetNames = theFile.sheetnames

print("All sheet names {} ".format(theFile.sheetnames))

currentSheet = theFile["Sheet1"]


# print(currentSheet['B4'].value)

# print max numbers of wors and colums for each sheet
# print(currentSheet.max_row)
# print(currentSheet.max_column)

def find_specific_cell():
    student_name = str(input("Enter student name: ")).strip().lower()
    for row in range(1, currentSheet.max_row + 1):
        cell_name = "{}{}".format("D", row)
        if currentSheet[cell_name].value.strip().lower() == student_name:
            # print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
            print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
            return cell_name


def get_column_letter(specificCellLetter):
    letter = specificCellLetter[0:-1]
    print(f'letter {letter}')
    return letter


def get_row_letter(specificCellLetter):
    letter = specificCellLetter[1:]
    print(f'letter {letter}')
    return letter


def get_all_values_by_cell_letter(letter):
    for row in range(1, currentSheet.max_row + 1):
        for column in letter:
            cell_name = "{}{}".format(column, row)
            # print(cell_name)
            print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))


def get_all_value_by_col_letter(letter):
    for column in ['S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR']:
        cell_name = "{}{}".format(column, letter)
        header = "{}{}".format(column, 1)
        print("cell position {} {} has value: {}".format(currentSheet[header].value.replace("\n", " "), cell_name,
                                                         currentSheet[cell_name].value))
        data = "".join(e for e in str(input("Enter {}: ".format(currentSheet[header].value.replace("\n", " ")))).strip() if e.isalnum() or e.isspace())
        currentSheet[cell_name] = " ".join(letter.capitalize() for letter in data.split(" "))
        theFile.save(url)


specificCellLetter = (find_specific_cell())
letter = get_row_letter(specificCellLetter)

get_all_value_by_col_letter(letter)
